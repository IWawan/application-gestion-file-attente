from datetime import datetime
import openpyxl
from typing import List

class extract_xlsx:
    def __init__(self, file_path: str):
        self.file_path = file_path

    def to_array(self) -> List[str]:
        array = []

        try:
            # Ouvrir le fichier Excel
            wb = openpyxl.load_workbook(self.file_path)
            sheet = wb.active  # Par défaut, prendre la première feuille
            nb_ligne = sheet.max_row

            # Parcourir les lignes du fichier, en partant de la ligne 2
            for i in range(2, nb_ligne + 1):
                heure_cell = sheet.cell(row=i, column=1)
                name_cell = sheet.cell(row=i, column=2)

                # Vérifier si la cellule "Usager" n'est pas vide
                if name_cell.value:
                    # Extraire l'heure
                    heure = heure_cell.value
                    if heure:
                        dt = datetime.strptime(heure, "%Y-%m-%d %H:%M:%S")
                        heure = dt.strftime("%H:%M")
                    else:
                        heure = ""

                    # Extraire le nom
                    name = name_cell.value

                    # Ajouter l'heure et le nom dans le tableau
                    array.append(f"{heure} | {name}")
            
            wb.close()  # Fermer le fichier Excel
        except Exception as e:
            print(f"Erreur d'extraction : {e}")

        return array
